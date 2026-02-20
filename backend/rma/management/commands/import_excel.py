"""
Management command to import RMA data from Excel file.
"""
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rma.models import RMA, RMAGroup
from datetime import datetime
from collections import defaultdict

User = get_user_model()


class Command(BaseCommand):
    help = 'Import RMA data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Path to Excel file'
        )
        parser.add_argument(
            '--admin-username',
            type=str,
            default='admin',
            help='Username of admin user to assign RMAs to'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        admin_username = options['admin_username']
        
        # Get admin user
        try:
            admin_user = User.objects.get(username=admin_username, role='ADMIN')
        except User.DoesNotExist:
            self.stdout.write(self.style.ERROR(
                f'Admin user "{admin_username}" not found. Please create an admin user first.'
            ))
            return
        
        self.stdout.write(f'Loading workbook from {excel_file}...')
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Failed to load Excel file: {e}'))
            return
        
        # Parse header row to find column indices
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_map = {cell: idx for idx, cell in enumerate(header_row) if cell}
        
        self.stdout.write(f'Found columns: {list(col_map.keys())}')
        
        # Group RMAs by month of first ship date
        rmas_by_month = defaultdict(list)
        
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            
            # Skip empty rows
            if not any(row):
                continue
            
            try:
                # Extract data from row
                serial_number = self._get_cell_value(row, col_map, 'SN', 'Serial Number')
                if not serial_number:
                    self.stdout.write(self.style.WARNING(f'Skipping row {row_count}: No serial number'))
                    continue
                
                # Parse dates - try multiple column name variations
                first_ship_date = self._parse_date(
                    self._get_cell_value(row, col_map, 'First Ship Date', '1st Ship Date', 'First Ship', 'First Shipped')
                )
                rma_received_date = self._parse_date(
                    self._get_cell_value(row, col_map, 'RMA Received Date', 'RMA rcvd', 'RMA Rcvd', 'Received Date')
                )
                return_date = self._parse_date(
                    self._get_cell_value(row, col_map, 'Return Date', 'Return date', 'Returned')
                )
                
                # Get RMA number from spreadsheet if available
                rma_number_from_sheet = self._get_cell_value(row, col_map, 'RMA', 'RMA Number', 'RMA #')
                
                # Determine group key (year-month of RMA received date preferred, then first ship)
                group_key = None
                if rma_received_date:
                    group_key = rma_received_date.strftime('%Y-%m')
                elif first_ship_date:
                    group_key = first_ship_date.strftime('%Y-%m')
                elif return_date:
                    group_key = return_date.strftime('%Y-%m')
                else:
                    group_key = 'unknown'
                
                # Determine state intelligently based on all available data
                state = self._determine_state(row, col_map, return_date, rma_received_date)
                
                # Determine priority based on notes/root cause severity
                priority = self._determine_priority(row, col_map)
                
                # Get all available data with multiple column name variations
                root_cause = self._get_cell_value(row, col_map, 'Root Cause', 'Root cause') or ''
                parts_replaced = self._get_cell_value(row, col_map, 'Part(s) Replaced', 'Parts Replaced', 'Parts') or ''
                cost_to_repair = self._get_cell_value(row, col_map, 'Cost to Repair', 'Cost', 'Repair Cost') or ''
                fault_notes = self._get_cell_value(row, col_map, 'Fault / Notes', 'Fault/Notes', 'Notes', 'Fault') or ''
                tx2_mac = self._get_cell_value(row, col_map, 'TX2 Mac', 'TX2 Mac Address', 'MAC Address', 'MAC') or ''
                rma_history = self._get_cell_value(row, col_map, 'RMA History?', 'RMA History', 'History') or ''
                years_in_field_calc = self._get_cell_value(row, col_map, 'Yrs in field', 'Years in Field', 'Years')
                
                # Parse boolean fields with multiple column variations
                script_ran = self._parse_bool(self._get_cell_value(row, col_map, 'Script ran?', 'Script ran', 'Script Ran'))
                services_enabled = self._parse_bool(self._get_cell_value(row, col_map, 'Services enabled?', 'Services enabled', 'Services Enabled'))
                uptime_good = self._parse_bool(self._get_cell_value(row, col_map, 'Uptime good?', 'Uptime good', 'Uptime Good'))
                stream_good = self._parse_bool(self._get_cell_value(row, col_map, 'Stream good?', 'Stream good', 'Stream Good'))
                ship_ready = self._parse_bool(self._get_cell_value(row, col_map, 'Ship ready', 'Ship Ready'))
                
                # Build comprehensive fault_notes including RMA history if present
                comprehensive_notes = fault_notes
                if rma_history:
                    if comprehensive_notes:
                        comprehensive_notes += f'\n\nRMA History: {rma_history}'
                    else:
                        comprehensive_notes = f'RMA History: {rma_history}'
                
                # Add spreadsheet years in field to notes if available
                if years_in_field_calc:
                    calc_note = f'\n[Imported: Years in field from spreadsheet: {years_in_field_calc}]'
                    comprehensive_notes += calc_note
                
                rma_data = {
                    'serial_number': str(serial_number),
                    'first_ship_date': first_ship_date,
                    'rma_received_date': rma_received_date,
                    'return_date': return_date,
                    'root_cause': root_cause,
                    'parts_replaced': parts_replaced,
                    'cost_to_repair': str(cost_to_repair) if cost_to_repair else '',
                    'fault_notes': comprehensive_notes or 'No notes available',
                    'tx2_mac': str(tx2_mac),
                    'script_ran': script_ran,
                    'services_enabled': services_enabled,
                    'uptime_good': uptime_good,
                    'stream_good': stream_good,
                    'ship_ready': ship_ready,
                    'state': state,
                    'priority': priority,
                    'owner': admin_user,
                }
                
                rmas_by_month[group_key].append(rma_data)
                
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'Error processing row {row_count}: {e}'))
                continue
        
        # Create RMA groups and RMAs
        total_created = 0
        for month_key, rma_list in rmas_by_month.items():
            if not rma_list:
                continue
            
            # Create group
            group = RMAGroup.objects.create(created_by=admin_user)
            
            self.stdout.write(f'\nCreating group for {month_key} with {len(rma_list)} RMAs...')
            
            for rma_data in rma_list:
                try:
                    rma = RMA.objects.create(group=group, **rma_data)
                    total_created += 1
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'Failed to create RMA for {rma_data["serial_number"]}: {e}'))
                    continue
        
        self.stdout.write(self.style.SUCCESS(
            f'\n✓ Successfully imported {total_created} RMAs in {len(rmas_by_month)} groups'
        ))
    
    def _get_cell_value(self, row, col_map, *possible_names):
        """Get cell value by trying multiple possible column names."""
        for name in possible_names:
            if name in col_map:
                idx = col_map[name]
                if idx < len(row):
                    return row[idx]
        return None
    
    def _parse_date(self, value):
        """Parse date from various formats."""
        if not value:
            return None
        
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, str):
            # Try various date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y', '%Y%m%d']:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        
        return None
    
    def _parse_bool(self, value):
        """Parse boolean value."""
        if value is None:
            return False
        
        if isinstance(value, bool):
            return value
        
        if isinstance(value, str):
            return value.lower() in ['yes', 'true', '1', 'x', 'checked']
        
        return bool(value)
    
    def _determine_state(self, row, col_map, return_date, rma_received_date):
        """Determine RMA state based on available data - intelligently."""
        # If return date is filled, it's completed (terminal state)
        if return_date:
            return 'COMPLETED'
        
        # Check if ship ready is checked - means it's been repaired and ready to ship
        ship_ready = self._parse_bool(self._get_cell_value(row, col_map, 'Ship ready', 'Ship Ready'))
        if ship_ready:
            return 'REPAIRED'
        
        # If parts replaced or cost to repair is filled, it's been repaired or replaced
        parts = self._get_cell_value(row, col_map, 'Part(s) Replaced', 'Parts Replaced', 'Parts')
        if parts and str(parts).strip():
            # Check if it says "replaced" vs individual parts
            if 'replace' in str(parts).lower() and 'unit' in str(parts).lower():
                return 'REPLACED'
            return 'REPAIRED'
        
        # If root cause is filled, it's been diagnosed
        root_cause = self._get_cell_value(row, col_map, 'Root Cause', 'Root cause')
        if root_cause and str(root_cause).strip():
            return 'DIAGNOSED'
        
        # If RMA received date is filled, it's been received
        if rma_received_date:
            return 'RECEIVED'
        
        # Default to approved (since these are historical RMAs that made it in)
        return 'APPROVED'
    
    def _determine_priority(self, row, col_map):
        """Determine priority based on notes and root cause."""
        # Get notes and root cause to check for severity indicators
        notes = str(self._get_cell_value(row, col_map, 'Fault / Notes', 'Fault/Notes', 'Notes', 'Fault') or '')
        root_cause = str(self._get_cell_value(row, col_map, 'Root Cause', 'Root cause') or '')
        
        combined = (notes + ' ' + root_cause).lower()
        
        # High priority indicators
        high_indicators = ['critical', 'urgent', 'broken', 'dead', 'failure', 'failed', 'not working', 'down']
        if any(indicator in combined for indicator in high_indicators):
            return 'HIGH'
        
        # Low priority indicators
        low_indicators = ['cosmetic', 'minor', 'aesthetic', 'scratch', 'dent']
        if any(indicator in combined for indicator in low_indicators):
            return 'LOW'
        
        # Default to normal
        return 'NORMAL'
